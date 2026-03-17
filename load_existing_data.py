"""
기존 통합공정진행표(20251126).xlsx 데이터를 대시보드용 data.json으로 변환하는 스크립트
한 번만 실행하면 됩니다.
"""

import openpyxl
import json
import os
from datetime import datetime

XLSX_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "통합공정진행표(20251126).xlsx")
DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.json")


def load_xlsx():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    daily_logs = []

    # --- 총괄표에서 일별 데이터 추출 ---
    ws = wb["총괄표"]

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        vals = [c.value for c in row]
        dt = vals[0]
        if not isinstance(dt, datetime):
            continue

        date_str = dt.strftime("%Y-%m-%d")

        def safe(v):
            return int(v) if isinstance(v, (int, float)) and v != 0 else 0

        # 각 공정별 데이터 추출 (총괄표 기준으로 "전체" 작업자로 기록)
        mappings = [
            ("분류", safe(vals[1]), safe(vals[3])),       # B=권, D=건
            ("면표시", safe(vals[7]), safe(vals[10])),     # H=권, K=면
            ("문서스캔", safe(vals[14]), safe(vals[17])),  # O=권, R=면
            ("도면스캔", safe(vals[21]), safe(vals[24])),  # V=권, Y=면
            ("보정", safe(vals[28]), safe(vals[31])),      # AC=권, AF=면
            ("색인", safe(vals[35]), safe(vals[38])),      # AJ=권, AM=면
            ("재편철", safe(vals[42]), safe(vals[43])),    # AQ=권호수, AR=건
            ("공개구분", safe(vals[48]), safe(vals[49])),  # AW=권호수, AX=건
        ]

        for proc, p_qty, s_qty in mappings:
            if p_qty > 0 or s_qty > 0:
                daily_logs.append({
                    "date": date_str,
                    "worker": "전체",
                    "process": proc,
                    "primary_qty": p_qty,
                    "secondary_qty": s_qty,
                })

    # --- 개별 시트에서 작업자별 데이터 추출 ---
    sheet_mappings = {
        "분류": {"sheet": "분류", "label": 0, "qty_col": 1, "worker_col": 2, "date_col": 3, "kwon_col": 5},
        "면표시": {"sheet": "면표시", "label": 0, "qty_col": 1, "worker_col": 2, "date_col": 3},
        "문서스캔": {"sheet": "문서스캔", "label": 0, "qty_col": 1, "worker_col": 2, "date_col": 3},
        "도면스캔": {"sheet": "도면스캔", "label": 0, "qty_col": 1, "worker_col": 2, "date_col": 3},
        "보정": {"sheet": "보정", "label": 0, "qty_col": 1, "worker_col": 2, "date_col": 3},
        "색인": {"sheet": "색인", "label": 0, "qty_col": 1, "worker_col": 2, "date_col": 3},
        "재편철": {"sheet": "재편철", "label": 0, "qty_col": 4, "worker_col": 1, "date_col": 2},
        "공개구분": {"sheet": "공개구분", "label": 0, "qty_col": 1, "worker_col": 2, "date_col": 3},
    }

    # 작업자별 일별 집계
    worker_daily = {}  # (date, worker, process) -> {primary, secondary}

    for proc, cfg in sheet_mappings.items():
        ws = wb[cfg["sheet"]]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[cfg["label"]] is None:
                continue
            worker = row[cfg["worker_col"]]
            dt = row[cfg["date_col"]]
            if not isinstance(dt, datetime) or not worker:
                continue

            date_str = dt.strftime("%Y-%m-%d")
            key = (date_str, str(worker), proc)

            if key not in worker_daily:
                worker_daily[key] = {"primary": 0, "secondary": 0}

            worker_daily[key]["primary"] += 1  # 레이블 1건 = 1권
            qty = row[cfg["qty_col"]]
            if isinstance(qty, (int, float)):
                worker_daily[key]["secondary"] += int(qty)

    # 작업자별 로그 생성
    worker_logs = []
    for (date_str, worker, proc), vals in worker_daily.items():
        worker_logs.append({
            "date": date_str,
            "worker": worker,
            "process": proc,
            "primary_qty": vals["primary"],
            "secondary_qty": vals["secondary"],
        })

    # 작업자 목록 추출
    workers = sorted(set(log["worker"] for log in worker_logs if log["worker"] != "전체"))

    # 기존 사업의 누적 총계 기반으로 목표 설정
    data = {
        "project": {
            "name": "2025년 중요기록물 정리사업",
            "start_date": "2025-04-01",
            "end_date": "2025-12-16",
            "total_kwon": 9897,
            "total_gun": 99107,
            "total_myun": 1530632,
        },
        "targets": {
            "분류": {"primary": 9897, "secondary": 99107},
            "면표시": {"primary": 9897, "secondary": 1530632},
            "문서스캔": {"primary": 9897, "secondary": 1530632},
            "도면스캔": {"primary": 1500, "secondary": 20000},
            "보정": {"primary": 9897, "secondary": 1530632},
            "색인": {"primary": 9897, "secondary": 1530632},
            "재편철": {"primary": 16542, "secondary": 99107},
            "공개구분": {"primary": 16542, "secondary": 99107},
        },
        "workers": workers if workers else [
            "정민혁", "정성원", "홍미정", "이원길",
            "윤경란", "김정의", "임미숙", "박미연",
        ],
        "daily_logs": worker_logs,
        "sampling_logs": [],
    }

    return data


def main():
    print("기존 xlsx 데이터를 로드 중...")
    data = load_xlsx()
    print(f"  - 일별 로그: {len(data['daily_logs'])}건")
    print(f"  - 작업자: {len(data['workers'])}명")
    print(f"  - 작업자 목록: {', '.join(data['workers'][:10])}...")

    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"\n✅ {DATA_FILE} 저장 완료!")
    print("이제 'streamlit run app.py' 로 대시보드를 실행하세요.")


if __name__ == "__main__":
    main()
